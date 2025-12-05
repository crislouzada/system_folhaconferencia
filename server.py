#!/usr/bin/env python3
"""
═══════════════════════════════════════════════════════════════════════════════
SERVIDOR FLASK - PROCESSAMENTO ESPECIALIZADO DE FOLHA DE PAGAMENTO
═══════════════════════════════════════════════════════════════════════════════

ESPECIALISTA EM:
✓ Reconhecimento de padrões de folha (ID - NOME DO FUNCIONÁRIO)
✓ Conversão precisa de valores (moeda, horas, decimais)
✓ Identificação automática de referências/competências
✓ Estruturação transposta para análise comparativa
✓ Detecção de divergências (Calculado vs Informado)

VERSÃO: 3.0 - Payroll Processing Engine
═══════════════════════════════════════════════════════════════════════════════
"""

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
import pandas as pd
import numpy as np
import re
import os
import tempfile
import traceback
from typing import Dict, List, Any, Optional
from decimal import Decimal, InvalidOperation

APP_VERSION = os.getenv('APP_VERSION', '3.0.1-functional')
app = Flask(__name__)
CORS(app)

# ═══════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÕES
# ═══════════════════════════════════════════════════════════════════════════

MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB

# Padrões de reconhecimento
EMPLOYEE_PATTERN = re.compile(r'^(\d+)\s*-\s*(.+)$')
REFERENCE_PATTERN = re.compile(r'(\d{1,2}/\d{4})')
TOTAL_PATTERN = re.compile(r'total', re.IGNORECASE)

# ═══════════════════════════════════════════════════════════════════════════
# FUNÇÕES DE CONVERSÃO DE VALORES (ESPECIALISTA)
# ═══════════════════════════════════════════════════════════════════════════

def parse_decimal_value(value: Any) -> float:
    """
    CONVERSÃO ROBUSTA DE VALORES - ESPECIALISTA EM FOLHA DE PAGAMENTO
    
    Suporta:
    1. Moeda BR: "4.077,32" → 4077.32
    2. Moeda US: "4,077.32" → 4077.32  
    3. Horas: "220:00" → 220.0 (decimal)
    4. Horas com minutos: "100:30" → 100.5
    5. Percentual: "12,5%" → 12.5
    6. Valores simples: "626,63" → 626.63
    
    Regra: Se não conseguir converter, retorna 0.0 (não quebra o fluxo)
    """
    
    if value is None or value == '' or pd.isna(value):
        return 0.0
    
    # Se já é número
    if isinstance(value, (int, float)):
        return float(value)
    
    # Converter para string e limpar
    str_value = str(value).strip()
    
    if not str_value or str_value == '-':
        return 0.0
    
    # PADRÃO 1: Horas (220:00, 100:30)
    # "220:00" → 220.0
    # "100:30" → 100.5 (100 horas e 30 minutos)
    if ':' in str_value:
        try:
            parts = str_value.split(':')
            hours = float(parts[0])
            minutes = float(parts[1]) if len(parts) > 1 else 0
            result = hours + (minutes / 60.0)
            print(f'    🕒 Convertido hora: {str_value} → {result:.2f}h')
            return result
        except Exception as e:
            print(f'    ⚠️  Erro ao converter hora "{str_value}": {e}')
            return 0.0
    
    # PADRÃO 2: Percentual (12,5%)
    if '%' in str_value:
        str_value = str_value.replace('%', '').strip()
    
    # Remover símbolos de moeda e espaços
    str_value = str_value.replace('R$', '').replace('$', '').replace(' ', '').replace('\xa0', '')
    
    # PADRÃO 3: Formato brasileiro com ponto e vírgula (4.077,32)
    if ',' in str_value and '.' in str_value:
        # Se ponto antes da vírgula: formato BR
        if str_value.rfind('.') < str_value.rfind(','):
            str_value = str_value.replace('.', '').replace(',', '.')
        # Se vírgula antes do ponto: formato US
        else:
            str_value = str_value.replace(',', '')
    
    # PADRÃO 4: Apenas vírgula (formato BR: 626,63)
    elif ',' in str_value:
        str_value = str_value.replace(',', '.')
    
    # PADRÃO 5: Apenas ponto (pode ser milhar ou decimal)
    elif '.' in str_value:
        parts = str_value.split('.')
        # Se tem 2 dígitos após o ponto, é decimal
        if len(parts) == 2 and len(parts[1]) <= 2:
            pass  # É decimal, manter
        # Múltiplos pontos = separador de milhar
        elif len(parts) > 2:
            str_value = str_value.replace('.', '')
        # Mais de 2 dígitos após ponto = milhar (ex: 1.000)
        elif len(parts) == 2 and len(parts[1]) > 2:
            str_value = str_value.replace('.', '')
    
    # Converter para float
    try:
        result = float(str_value)
        return result
    except Exception as e:
        print(f'    ⚠️  Não foi possível converter "{value}" → retornando 0.0')
        return 0.0


def detect_value_type(value: Any) -> str:
    """
    Detecta tipo de valor:
    - 'currency': Moeda (valores >= 10.0)
    - 'hours': Horas (formato HH:MM ou decimal < 1000)
    - 'percentage': Percentual
    - 'integer': Inteiro
    """
    
    str_val = str(value).strip()
    
    if ':' in str_val:
        return 'hours'
    
    if '%' in str_val:
        return 'percentage'
    
    numeric = parse_decimal_value(value)
    
    if numeric >= 10.0:
        return 'currency'
    elif 0 < numeric < 10 and numeric % 1 != 0:
        return 'hours'
    else:
        return 'integer'


# ═══════════════════════════════════════════════════════════════════════════
# ESTRUTURAÇÃO INTELIGENTE DE DADOS
# ═══════════════════════════════════════════════════════════════════════════

def structure_payroll_data(raw_data: List[List[str]]) -> Dict[str, Any]:
    """
    ESTRUTURAÇÃO INTELIGENTE DE DADOS DE FOLHA DE PAGAMENTO
    
    ENTRADA: Lista de listas (tabela Excel)
    
    LÓGICA:
    1. Detecta funcionários pelo padrão: "NÚMERO - NOME" (ex: "7 - ALEX BARBOZA DE MELO")
    2. Para cada funcionário, coleta eventos até o próximo funcionário ou fim do arquivo
    3. Agrupa eventos por código+descrição e referência
    4. Monta estrutura TRANSPOSTA: uma linha por evento, colunas por referência
    
    SAÍDA: {
        'employees': [
            {
                'id': '7',
                'name': 'ALEX BARBOZA DE MELO',
                'references': ['10/2025', '11/2025'],
                'events': [
                    {
                        'code': '1',
                        'description': 'HORAS NORMAIS',
                        'values': {
                            '10/2025': {'calculated': 4077.32, 'informed': 220.0, 'difference': 3857.32},
                            '11/2025': {'calculated': 4362.73, 'informed': 220.0, 'difference': 4142.73}
                        }
                    }
                ]
            }
        ],
        'allReferences': ['10/2025', '11/2025']
    }
    """
    
    if not raw_data or len(raw_data) == 0:
        return {'employees': [], 'allReferences': [], 'summary': {}, 'companyInfo': {}}
    
    employees = []
    all_references = set()
    current_employee = None
    company_info = {}
    
    print('\n' + '═' * 80)
    print('📊 ESTRUTURANDO DADOS DE FOLHA DE PAGAMENTO')
    print('═' * 80)
    print(f'📋 Total de linhas: {len(raw_data)}')
    
    # Extrair informações da empresa das primeiras linhas
    for idx, row in enumerate(raw_data[:5]):
        row_str = ' '.join([str(cell) for cell in row if str(cell).strip() and str(cell).strip() != 'nan'])
        
        # Buscar Empresa
        if 'Empresa:' in row_str or 'empresa:' in row_str.lower():
            for i, cell in enumerate(row):
                cell_str = str(cell).strip()
                if ' - ' in cell_str and 'Empresa' not in cell_str:
                    company_info['name'] = cell_str
                    break
        
        # Buscar CNPJ
        if 'CNPJ:' in row_str or 'cnpj:' in row_str.lower():
            for i, cell in enumerate(row):
                cell_str = str(cell).strip()
                if '/' in cell_str and '-' in cell_str and len(cell_str) >= 14:
                    company_info['cnpj'] = cell_str
                    break
        
        # Buscar Competência
        if 'Competência:' in row_str or 'competencia:' in row_str.lower():
            refs = []
            for i, cell in enumerate(row):
                cell_str = str(cell).strip()
                if REFERENCE_PATTERN.search(cell_str):
                    refs.append(cell_str)
            if refs:
                company_info['period'] = ' até '.join(refs) if len(refs) > 1 else refs[0]
    
    if company_info:
        print(f'\n📄 INFORMAÇÕES DA EMPRESA:')
        if 'name' in company_info:
            print(f'   🏢 Empresa: {company_info["name"]}')
        if 'cnpj' in company_info:
            print(f'   📋 CNPJ: {company_info["cnpj"]}')
        if 'period' in company_info:
            print(f'   📅 Período: {company_info["period"]}')
    
    # DEBUG: Mostrar primeiras linhas para entender estrutura
    print('\n🔍 DEBUG - Primeiras 15 linhas do arquivo:')
    for idx, row in enumerate(raw_data[:15]):
        # Mostrar apenas colunas não vazias
        non_empty = [(i, str(cell)[:30]) for i, cell in enumerate(row) if str(cell).strip() and str(cell).strip() != 'nan']
        if non_empty:
            print(f'   Linha {idx}: {non_empty}')
    
    # Detectar índices de colunas (buscar nas primeiras 10 linhas)
    col_indices = {}
    for row in raw_data[:10]:
        temp_indices = detect_column_indices(row)
        # Se encontrou pelo menos 3 colunas identificadas, usar esse mapeamento
        if len([v for v in temp_indices.values() if v >= 0]) >= 3:
            col_indices = temp_indices
            print(f'\n🗺️  Colunas detectadas na linha: {row[:5]}...')
            break
    
    # Se não encontrou, usar posições padrão
    if not col_indices:
        col_indices = {'code': 0, 'description': 1, 'reference': 2, 'calculated': 3, 'informed': 4}
        print(f'⚠️  Usando mapeamento padrão de colunas')
    
    print(f'🗺️  Mapeamento final: {col_indices}\n')
    
    # Processar linhas
    event_count = 0
    employees_map = {}  # Para consolidar funcionários duplicados pelo ID
    
    for row_idx, row in enumerate(raw_data[1:], start=2):
        
        if not row or all(str(cell).strip() == '' for cell in row):
            continue
        
        # DETECÇÃO DE FUNCIONÁRIO: Buscar padrão "NÚMERO - NOME" em qualquer coluna
        employee_found = False
        for col_idx, cell in enumerate(row):
            cell_str = str(cell).strip()
            
            # Regex: ^(\d+)\s*-\s*(.+)$
            match = EMPLOYEE_PATTERN.match(cell_str)
            
            if match:
                # Novo funcionário detectado
                emp_id = match.group(1).strip()
                emp_name = match.group(2).strip()
                
                # FILTRO: Ignorar se for nome de empresa (contém LTDA, ME, EPP, etc)
                company_keywords = ['LTDA', 'ME', 'EPP', 'EIRELI', 'S.A', 'S/A', 'CIA']
                if any(keyword in emp_name.upper() for keyword in company_keywords):
                    print(f'   ⏭️  Empresa ignorada: {emp_id} - {emp_name}')
                    current_employee = None  # Resetar para não processar eventos da empresa
                    employee_found = True
                    break
                
                # Verificar se funcionário já existe (duplicado por quebra de página)
                if emp_id in employees_map:
                    # Reativar funcionário existente
                    current_employee = employees_map[emp_id]
                    print(f'\n🔄 Funcionário duplicado detectado: {emp_id} - {emp_name} (consolidando eventos)')
                else:
                    # Salvar funcionário anterior se existir e for diferente
                    if current_employee and current_employee['id'] not in employees_map:
                        current_employee['references'] = sorted(list(current_employee['references']))
                        current_employee['events'] = convert_to_transposed_structure(
                            current_employee['events_map'], 
                            current_employee['references']
                        )
                        current_employee['totals'] = calculate_employee_totals(
                            current_employee['events'], 
                            current_employee['references']
                        )
                        del current_employee['events_map']
                    
                    # Criar novo funcionário
                    current_employee = {
                        'id': emp_id,
                        'name': emp_name,
                        'events_map': {},  # {event_key: {ref: {calc, info}}}
                        'references': set()
                    }
                    
                    employees_map[emp_id] = current_employee
                    employees.append(current_employee)
                    
                    print(f'\n👤 Funcionário #{len(employees)}: {emp_id} - {emp_name}')
                
                employee_found = True
                break
        
        # Se é linha de funcionário, pular para próxima linha
        if employee_found:
            continue
        
        # Se temos funcionário atual, processar linha como evento
        if current_employee:
            
            # Verificar se é linha de total (ignorar)
            first_col = str(row[0]).lower() if row else ''
            if TOTAL_PATTERN.search(first_col):
                print(f'   ⏭️  Linha de total ignorada: {first_col}')
                continue
            
            # Extrair dados do evento
            code = str(row[col_indices.get('code', 0)]).strip() if col_indices.get('code', 0) < len(row) else ''
            description = str(row[col_indices.get('description', 1)]).strip() if col_indices.get('description', 1) < len(row) else ''
            reference = str(row[col_indices.get('reference', 2)]).strip() if col_indices.get('reference', 2) < len(row) else ''
            calculated_raw = row[col_indices.get('calculated', 3)] if col_indices.get('calculated', 3) < len(row) else 0
            informed_raw = row[col_indices.get('informed', 4)] if col_indices.get('informed', 4) < len(row) else 0
            tipo_raw = None
            if 'type' in col_indices and col_indices['type'] < len(row):
                tipo_raw = row[col_indices['type']]
            
            # Validar dados essenciais
            if not code or not reference:
                continue
            
            # Limpar referência usando regex
            ref_match = REFERENCE_PATTERN.search(reference)
            if ref_match:
                reference = ref_match.group(1)
            else:
                # Se não encontrou padrão de referência, pular
                continue
            
            # Converter valores com função robusta
            calculated = parse_decimal_value(calculated_raw)
            informed = parse_decimal_value(informed_raw)

            # Aplicar regra de sinal baseada no TIPO em ambos os campos:
            # 'P' (ou 'p') = positivo; diferente de 'P' = negativo
            if tipo_raw is not None:
                tipo_flag = str(tipo_raw).strip().upper()[:1]
                sign = 1 if tipo_flag == 'P' else -1
                calculated = abs(calculated) * sign
                informed = abs(informed) * sign

            # Padronizar para 2 casas decimais
            calculated = round(calculated, 2)
            informed = round(informed, 2)
            
            # Adicionar referência aos sets
            all_references.add(reference)
            current_employee['references'].add(reference)
            
            # Chave única do evento (código + descrição)
            event_key = f"{code}|||{description}"
            
            # Criar estrutura se não existe
            if event_key not in current_employee['events_map']:
                current_employee['events_map'][event_key] = {}
            
            # Armazenar valores por referência
            current_employee['events_map'][event_key][reference] = {
                'calculated': calculated,
                'informed': informed,
                'difference': round(calculated - informed, 2)
            }
            
            event_count += 1
            
            if event_count <= 5:  # Mostrar apenas os primeiros 5 eventos por funcionário
                print(f'   📝 {code} - {description[:40]:40s} | {reference} | Calc: {calculated:>10.2f} | Info: {informed:>10.2f}')
    
    # Processar todos os funcionários consolidados
    for emp in employees:
        if 'events_map' in emp:
            emp['references'] = sorted(list(emp['references']))
            emp['events'] = convert_to_transposed_structure(
                emp['events_map'], 
                emp['references']
            )
            emp['totals'] = calculate_employee_totals(
                emp['events'], 
                emp['references']
            )
            del emp['events_map']
    
    # Ordenar referências globalmente
    sorted_references = sorted(list(all_references))
    
    print(f'\n✅ Estruturação completa:')
    print(f'   👥 {len(employees)} funcionários únicos')
    print(f'   📅 {len(sorted_references)} referências: {sorted_references}')
    print(f'   📊 Total de eventos processados: {event_count}')
    
    return {
        'employees': employees,
        'allReferences': sorted_references,
        'summary': calculate_global_summary(employees, sorted_references),
        'companyInfo': company_info
    }


def detect_column_indices(headers: List[str]) -> Dict[str, int]:
    """
    Detecta índices de colunas importantes baseado em padrões
    """
    
    col_map = {}
    
    patterns = {
        'code': ['codigo', 'código', 'cod', 'cód'],
        'description': ['nome', 'descrição', 'descricao', 'historico', 'descrição do evento'],
        'reference': ['referencia', 'referência', 'ref', 'competencia', 'competência'],
        'calculated': ['calculado', 'valor calculado', 'calc', 'vlr calc'],
        'informed': ['informado', 'valor informado', 'inf', 'vlr inf'],
        'type': ['tipo', 'tp', 'p/d', 'pd', 'natureza']
    }
    
    for col_idx, header in enumerate(headers):
        header_lower = str(header).lower().strip()
        
        for key, keywords in patterns.items():
            if any(kw in header_lower for kw in keywords):
                col_map[key] = col_idx
                break
    
    # Se não encontrou, usar posições padrão do formato Excel
    # Baseado na análise: Linha 9: [(0, '1'), (4, 'HORAS NORMAIS'), (17, '10/2025'), (20, '4.077,32'), (23, '220:00')]
    if 'code' not in col_map:
        col_map['code'] = 0
    if 'description' not in col_map:
        col_map['description'] = 4  # Mudou de 1 para 4
    if 'reference' not in col_map:
        col_map['reference'] = 17  # Mudou de 2 para 17
    if 'calculated' not in col_map:
        col_map['calculated'] = 20  # Mudou de 3 para 20
    if 'informed' not in col_map:
        col_map['informed'] = 23  # Mudou de 4 para 23
    # 'type' é opcional; não define padrão se não detectado
    
    return col_map


def convert_to_transposed_structure(events_map: Dict, references: List[str]) -> List[Dict]:
    """
    Converte Map de eventos para estrutura transposta
    """
    
    events = []
    
    for event_key, ref_values in events_map.items():
        code, description = event_key.split('|||')
        
        event = {
            'code': code,
            'description': description,
            'values': {}
        }
        
        # Para cada referência, adicionar valores
        for ref in references:
            if ref in ref_values:
                event['values'][ref] = ref_values[ref]
            else:
                event['values'][ref] = {
                    'calculated': 0.0,
                    'informed': 0.0,
                    'difference': 0.0
                }
        
        events.append(event)
    
    # Ordenar por código numérico
    events.sort(key=lambda e: int(e['code']) if e['code'].isdigit() else 9999)
    
    return events


def calculate_employee_totals(events: List[Dict], references: List[str]) -> Dict:
    """
    Calcula totais por referência para um funcionário
    """
    
    totals = {}
    
    for ref in references:
        calc_sum = sum(e['values'][ref]['calculated'] for e in events)
        info_sum = sum(e['values'][ref]['informed'] for e in events)
        calc_sum = round(calc_sum, 2)
        info_sum = round(info_sum, 2)
        diff_sum = round(calc_sum - info_sum, 2)

        totals[ref] = {
            'calculated': calc_sum,
            'informed': info_sum,
            'difference': diff_sum
        }
    
    return totals


def calculate_global_summary(employees: List[Dict], references: List[str]) -> Dict:
    """
    Calcula resumo global de todos os funcionários
    """
    
    summary = {
        'total_employees': len(employees),
        'total_events': sum(len(emp['events']) for emp in employees),
        'by_reference': {}
    }
    
    for ref in references:
        total_calc = sum(emp['totals'][ref]['calculated'] for emp in employees if ref in emp['totals'])
        total_info = sum(emp['totals'][ref]['informed'] for emp in employees if ref in emp['totals'])
        total_calc = round(total_calc, 2)
        total_info = round(total_info, 2)
        total_diff = round(total_calc - total_info, 2)

        summary['by_reference'][ref] = {
            'total_calculated': total_calc,
            'total_informed': total_info,
            'total_difference': total_diff
        }
    
    return summary


# ═══════════════════════════════════════════════════════════════════════════
# ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    """Serve a página principal"""
    return send_from_directory('.', 'index_v2.html')


@app.route('/index_v2.html')
def index_v2():
    """Serve a página principal"""
    return send_from_directory('.', 'index_v2.html')


@app.route('/app_v2.js')
def app_js():
    """Serve o JavaScript"""
    return send_from_directory('.', 'app_v2.js')


@app.route('/ajuda.html')
def ajuda():
    """Serve a página de ajuda"""
    return send_from_directory('.', 'ajuda.html')


@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy', 'version': APP_VERSION}), 200


@app.route('/parse-excel', methods=['POST'])
def parse_excel():
    """
    Endpoint principal - Processa arquivos de folha de pagamento
    """
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'errorCode': 'NO_FILE', 'message': 'Nenhum arquivo enviado'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'errorCode': 'NO_FILE', 'message': 'Nome de arquivo vazio'}), 400
    
    original_filename = secure_filename(file.filename)
    
    try:
        # Salvar temporariamente
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(original_filename)[1])
        temp_path = temp_file.name
        file.save(temp_path)
        
        file_size = os.path.getsize(temp_path)
        extension = os.path.splitext(original_filename)[1].lower()
        
        print('\n' + '═' * 80)
        print(f'📄 ARQUIVO: {original_filename}')
        print(f'💾 Tamanho: {file_size:,} bytes')
        print(f'📝 Extensão: {extension}')
        print('═' * 80)
        
        # Ler arquivo
        df = None
        
        if extension in ['.csv', '.txt']:
            encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
            for enc in encodings:
                try:
                    df = pd.read_csv(temp_path, encoding=enc, header=None, sep=None, engine='python')
                    print(f'✅ CSV lido com encoding: {enc}')
                    break
                except:
                    continue
        
        elif extension == '.xlsx':
            print('🔄 Tentando ler XLSX...')
            
            # ESTRATÉGIA 1: Tentar ler diretamente com openpyxl
            try:
                from openpyxl import load_workbook
                
                # Carregar workbook com openpyxl diretamente
                wb = load_workbook(filename=temp_path, read_only=True, data_only=True)
                print(f'  📑 Workbook carregado: {wb.sheetnames}')
                
                if len(wb.sheetnames) == 0:
                    return jsonify({
                        'success': False,
                        'errorCode': 'EMPTY_SHEETS',
                        'message': 'Arquivo XLSX sem planilhas',
                        'suggestion': '💡 Abra no Excel e salve como CSV UTF-8'
                    }), 400
                
                # Pegar primeira sheet ou a sheet "Movimentos"
                sheet_name = None
                if 'Movimentos' in wb.sheetnames:
                    sheet_name = 'Movimentos'
                else:
                    sheet_name = wb.sheetnames[0]
                
                print(f'  📄 Lendo sheet: {sheet_name}')
                
                # Ler com pandas usando a sheet específica
                df = pd.read_excel(temp_path, sheet_name=sheet_name, engine='openpyxl', header=None)
                print(f'  ✅ XLSX lido com sucesso: {df.shape[0]} linhas x {df.shape[1]} colunas')
                
                wb.close()
                
            except Exception as e1:
                error_msg = str(e1)
                print(f'  ⚠️  Erro: {error_msg[:200]}')
                
                # ESTRATÉGIA 2: Tentar sem especificar sheet
                try:
                    print('  🔄 Tentativa 2: leitura sem sheet específica')
                    df = pd.read_excel(temp_path, engine='openpyxl', header=None)
                    print(f'  ✅ Sucesso: {df.shape[0]} linhas x {df.shape[1]} colunas')
                    
                except Exception as e2:
                    print(f'  ❌ Falhou: {str(e2)[:200]}')
                    
                    # Se falhou tudo, pedir CSV
                    return jsonify({
                        'success': False,
                        'errorCode': 'XLSX_READ_ERROR',
                        'message': 'Não foi possível ler o arquivo XLSX',
                        'suggestion': '💡 SOLUÇÃO: No Excel, vá em Arquivo → Salvar Como → CSV UTF-8',
                        'details': f'Erro 1: {str(e1)[:100]} | Erro 2: {str(e2)[:100]}'
                    }), 400
        
        elif extension == '.xls':
            # Para XLS, tentar openpyxl primeiro (não precisa de xlrd 1.2.0)
            try:
                df = pd.read_excel(temp_path, engine='openpyxl', header=None)
                print(f'✅ XLS lido com openpyxl')
            except Exception as e1:
                # Se falhar, tentar sem engine (Pandas escolhe automaticamente)
                try:
                    df = pd.read_excel(temp_path, header=None)
                    print(f'✅ XLS lido com engine padrão')
                except Exception as e2:
                    return jsonify({
                        'success': False,
                        'errorCode': 'CORRUPTED_FILE',
                        'message': 'Arquivo XLS corrompido ou ilegível',
                        'suggestion': '💡 SOLUÇÃO: Abra no Excel e salve como CSV UTF-8',
                        'details': f'Tentativas falharam: {str(e1)[:100]} | {str(e2)[:100]}'
                    }), 400
        
        if df is None or df.empty:
            return jsonify({
                'success': False,
                'errorCode': 'PARSING_FAILED',
                'message': 'Não foi possível ler o arquivo'
            }), 400
        
        # Converter para lista
        raw_data = df.fillna('').astype(str).values.tolist()
        
        # Estruturar dados
        structured = structure_payroll_data(raw_data)
        
        print(f'\n✅ PROCESSAMENTO CONCLUÍDO')
        print(f'   👥 {structured["summary"]["total_employees"]} funcionários')
        print(f'   📝 {structured["summary"]["total_events"]} eventos')
        
        return jsonify({
            'success': True,
            'data': raw_data,
            'structured': structured,
            'filename': original_filename
        }), 200
        
    except Exception as e:
        print(f'\n❌ ERRO: {str(e)}')
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'errorCode': 'PROCESSING_ERROR',
            'message': str(e)
        }), 500
        
    finally:
        if 'temp_path' in locals() and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except:
                pass


# ═══════════════════════════════════════════════════════════════════════════
# EXECUÇÃO
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print('\n' + '═' * 80)
    print('🚀 SERVIDOR DE PROCESSAMENTO DE FOLHA DE PAGAMENTO V3.0')
    print('═' * 80)
    print('✓ Reconhecimento inteligente de padrões')
    print('✓ Conversão precisa de valores decimais')
    print('✓ Estruturação transposta para comparativo')
    print('✓ Detecção automática de referências')
    print('=' * 80)
    print(f'🌐 Versão: {APP_VERSION}')
    print(f'🌐 Servidor: http://localhost:5001')
    print(f'📡 Endpoint: POST /parse-excel')
    print('=' * 80 + '\n')
    
    app.run(host='0.0.0.0', port=5001, debug=True)
